"""
Title & Description Generator

Refactored to live under Agents/ and resolve all paths, sheet names, and column
labels from the central config module. Falls back to sensible defaults if a
constant is not present in config for backward compatibility.

Usage:
    python Agents/title_desc_generator.py

Requirements:
    - pandas
    - openpyxl

Notes:
    - Preserves workbook formatting by editing via openpyxl and only using
      pandas for row-wise transforms.
    - Creates the "Listings" sheet on first run if it is missing.
"""
from __future__ import annotations

from pathlib import Path
import sys
import math
from typing import Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Ensure the project root (where config.py lives) is in sys.path when running from Agents/
HERE = Path(__file__).resolve()
# Discover project root by walking up until we find config.py
CONFIG_BASENAME = "config.py"
PROJECT_ROOT = None
for _p in HERE.parents:
    if (_p / CONFIG_BASENAME).exists():
        PROJECT_ROOT = _p
        break
# Fallback to Agents/ -> project root assumption
if PROJECT_ROOT is None:
    PROJECT_ROOT = HERE.parent.parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    import config as cfg
except Exception as exc:
    raise RuntimeError(
        "Unable to import project config. Ensure config.py is at project root."
    ) from exc

# ---- Config lookups with safe fallbacks ------------------------------------
INVENTORY_XLSX = getattr(cfg, "INVENTORY_FILE_PATH", PROJECT_ROOT / "Inventory.xlsx")
INVENTORY_SHEET = getattr(cfg, "INVENTORY_SHEET_NAME", "Inventory")
# Many projects keep a Listings sheet separate from Inventory
LISTINGS_SHEET = getattr(cfg, "LISTINGS_SHEET_NAME", "Listings")

# Inventory column names
SKU_COL = getattr(cfg, "SKU_COLUMN", "SKU (Old)")
STATUS_COL = getattr(cfg, "STATUS_COLUMN", "Status")
BRAND_COL = getattr(cfg, "BRAND_COLUMN", "Brand")
KEYWORDS_COL = getattr(cfg, "KEYWORDS_COLUMN", "Keywords")
COLOR_COL = getattr(cfg, "COLOR_COLUMN", "Color")
SIZE_COL = getattr(cfg, "SIZE_COLUMN", "Size")
OP_COL = getattr(cfg, "OP_COLUMN", "OP")
CONDITION_COL = getattr(cfg, "CONDITION_COLUMN", "Condition")
EAN_COL = getattr(cfg, "EAN_COLUMN", "EAN")
MATERIALS_COL = getattr(cfg, "MATERIALS_COLUMN", "Materials")
MORE_DETAILS_COL = getattr(cfg, "MORE_DETAILS_COLUMN", "More details")

# Listings column names
LISTINGS_SKU_COL = getattr(cfg, "LISTINGS_SKU_COLUMN", SKU_COL)
LISTINGS_TITLE_COL = getattr(cfg, "LISTINGS_TITLE_COLUMN", "Title")
LISTINGS_DESC_COL = getattr(cfg, "LISTINGS_DESCRIPTION_COLUMN", "Text Description")

READY_VALUE = getattr(cfg, "READY_STATUS_VALUE", "Ready")
OK_VALUE = getattr(cfg, "OK_STATUS_VALUE", "OK")

# ---- Helpers ----------------------------------------------------------------

def _ws_to_dataframe(ws: Worksheet) -> pd.DataFrame:
    """Convert an openpyxl worksheet with headers in the first row to a DataFrame."""
    df = pd.DataFrame(ws.values)
    if df.empty:
        return pd.DataFrame()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    return df


def _ensure_listings_sheet(wb) -> Worksheet:
    if LISTINGS_SHEET in wb.sheetnames:
        return wb[LISTINGS_SHEET]
    ws = wb.create_sheet(LISTINGS_SHEET)
    # Initialize headers
    ws.append([LISTINGS_SKU_COL, LISTINGS_TITLE_COL, LISTINGS_DESC_COL])
    return ws


def _safe_get(row: pd.Series, col: str, default: str = "") -> str:
    val = row.get(col, default)
    # Treat NaN/None as empty string
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    return str(val)


# ---- Business logic ---------------------------------------------------------

def create_title(row: pd.Series) -> str:
    brand = _safe_get(row, BRAND_COL)
    keyword = _safe_get(row, KEYWORDS_COL)
    color = _safe_get(row, COLOR_COL)
    size = _safe_get(row, SIZE_COL)    # OP removed from title generation

    if not keyword:
        raise ValueError(f"Keyword is missing in row with index {row.name}")

    parts = []
    parts.append(f"{brand} {keyword}".strip() if brand else keyword)
    if color:
        parts.append(color)
    if size:
        parts.append(f"Größe {size}")

    title = ", ".join(parts)

    # If too long, drop the word "Größe"
    if len(title) > 80 and size:
        title = title.replace(f"Größe {size}", size)
    if len(title) > 80:
        raise ValueError(f"Title length exceeds 80 characters in row with index {row.name}")

    return title


def create_description(row: pd.Series) -> str:
    condition = _safe_get(row, CONDITION_COL)
    size = _safe_get(row, SIZE_COL)
    color = _safe_get(row, COLOR_COL)
    brand = _safe_get(row, BRAND_COL)
    ean = _safe_get(row, EAN_COL)
    materials = _safe_get(row, MATERIALS_COL)
    more_info = _safe_get(row, MORE_DETAILS_COL)

    # Remove decimal from numeric EANs read as floats
    try:
        if ean and float(ean).is_integer():
            ean = str(int(float(ean)))
    except ValueError:
        pass

    lines = []
    if condition:
        lines.append(f"Zustand: {condition}")
    if size:
        lines.append(f"Größe: {size}")
    if color:
        lines.append(f"Farbe: {color}")
    if brand:
        lines.append(f"Marke: {brand}")
    if ean:
        lines.append(f"EAN: {ean}")
    if more_info:
        lines.append("")
        lines.append(more_info)
    if materials:
        lines.append("")
        lines.append(materials)

    return "\n".join(lines)

# ---- Main -------------------------------------------------------------------

def main() -> None:
    print("Loading workbook...")
    wb = load_workbook(INVENTORY_XLSX, data_only=False)

    if INVENTORY_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{INVENTORY_SHEET}' not found in {INVENTORY_XLSX}")

    ws_inventory = wb[INVENTORY_SHEET]
    ws_listings = _ensure_listings_sheet(wb)

    print("Reading Inventory and Listings sheets...")
    df_inventory = _ws_to_dataframe(ws_inventory)
    df_listings = _ws_to_dataframe(ws_listings)

    if df_listings.empty or any(c not in df_listings.columns for c in [LISTINGS_SKU_COL, LISTINGS_TITLE_COL, LISTINGS_DESC_COL]):
        df_listings = pd.DataFrame(columns=[LISTINGS_SKU_COL, LISTINGS_TITLE_COL, LISTINGS_DESC_COL])

    print("Filtering rows with status = Ready...")
    df_ready = df_inventory[df_inventory[STATUS_COL] == READY_VALUE].copy()
    print(f"Found {len(df_ready)} rows to process.")

    results: Dict[Any, Dict[str, str]] = {}

    for idx, row in df_ready.iterrows():
        try:
            title = create_title(row)
            description = create_description(row)
            results[idx] = {"title": title, "description": description}
            df_inventory.loc[idx, STATUS_COL] = OK_VALUE
        except ValueError as e:
            print(f"Skipped row {idx}: {e}")

    if not results:
        print("No rows processed. Nothing to update.")
        wb.save(INVENTORY_XLSX)
        return

    print(f"Generating output for {len(results)} SKUs...")

    # ---- Write back to workbook --------------------------------------------
    # Write Inventory status back (header row is 1; data starts at row 2)
    # Map of column name -> 1-based column index on the sheet
    inv_header = {cell.value: i + 1 for i, cell in enumerate(next(ws_inventory.iter_rows(min_row=1, max_row=1))[0:len(df_inventory.columns)])}
    status_col_idx = inv_header.get(STATUS_COL)
    if status_col_idx is None:
        raise KeyError(f"Could not locate column '{STATUS_COL}' in Inventory header row")

    for i, status in enumerate(df_inventory[STATUS_COL].tolist(), start=2):
        ws_inventory.cell(row=i, column=status_col_idx, value=status)

    # ---- Non-destructive Listings update (preserve formulas/other columns) ----
    # Ensure header exists in sheet
    if ws_listings.max_row == 0:
        ws_listings.cell(row=1, column=1, value=LISTINGS_SKU_COL)
        ws_listings.cell(row=1, column=2, value=LISTINGS_TITLE_COL)
        ws_listings.cell(row=1, column=3, value=LISTINGS_DESC_COL)

    # Build header map for Listings (robust)
    lst_header = {}
    for c in range(1, ws_listings.max_column + 1):
        header_val = ws_listings.cell(row=1, column=c).value
        if header_val:
            lst_header[str(header_val).strip()] = c

    # Ensure required columns exist; if missing, add them at the end
    for required in [LISTINGS_SKU_COL, LISTINGS_TITLE_COL, LISTINGS_DESC_COL]:
        if required not in lst_header:
            new_c = ws_listings.max_column + 1
            ws_listings.cell(row=1, column=new_c, value=required)
            lst_header[required] = new_c

    sku_idx = lst_header[LISTINGS_SKU_COL]
    title_idx = lst_header[LISTINGS_TITLE_COL]
    desc_idx = lst_header[LISTINGS_DESC_COL]

    # Map existing SKUs -> row numbers
    existing_rows = {}
    for r in range(2, ws_listings.max_row + 1):
        v = ws_listings.cell(row=r, column=sku_idx).value
        if v is None or (isinstance(v, float) and math.isnan(v)):
            continue
        v_str = str(int(v)) if isinstance(v, float) and float(v).is_integer() else str(v)
        existing_rows[v_str] = r

    updated = 0
    created = 0

    # Write only the two target columns (plus SKU for new rows)
    for idx, payload in results.items():
        sku = df_ready.loc[idx, SKU_COL] if SKU_COL in df_ready.columns else None
        if sku is None or (isinstance(sku, float) and pd.isna(sku)):
            continue
        sku_str = str(int(sku)) if isinstance(sku, float) and float(sku).is_integer() else str(sku)

        if sku_str in existing_rows:
            r = existing_rows[sku_str]
            ws_listings.cell(row=r, column=title_idx, value=payload["title"])
            ws_listings.cell(row=r, column=desc_idx, value=payload["description"])
            updated += 1
        else:
            r = ws_listings.max_row + 1
            # Set only our three cells. No append of full row, so no other cells are touched.
            ws_listings.cell(row=r, column=sku_idx, value=sku_str)
            ws_listings.cell(row=r, column=title_idx, value=payload["title"])
            ws_listings.cell(row=r, column=desc_idx, value=payload["description"])
            created += 1

    wb.save(INVENTORY_XLSX)
    print(f"Done. Updated {updated} rows, created {created} new rows.")



if __name__ == "__main__":
    main()

