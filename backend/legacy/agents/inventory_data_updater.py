"""
Import per-SKU JSON files back into the inventory Excel sheet (table-aware updater).

- Reads ./products/<SKU>.json produced by `inventory_data_collector.py`.
- Flattens known categories/fields and updates only changed cells in the
  inventory sheet, preserving the Excel Table formatting by editing
  in-place with openpyxl (no re-writing the sheet via pandas).
- Skips overwriting when the incoming value is identical after normalization.
- Optionally appends missing SKUs (disabled by default; enable with --append-missing).

Usage:
  python inventory_json_importer.py [--products-dir ./products] [--append-missing]

Assumptions:
- Inventory workbook and sheet names, as well as column names, are defined in `config.py`.
- The sheet contains a header row with the defined column names.
- The SKU column exists and uniquely identifies rows.
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Resolve project root and import config and exporter constants
ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT.parent))  # allow importing project modules at repo root
import config  # type: ignore

JSON_COLUMN = getattr(config, "JSON_COLUMN", "JSON")

IMAGE_COUNT_SOURCES = [
    ("phone", getattr(config, "IMAGES_JSON_PHONE_COLUMN", "Images JSON Phone")),
    ("stock", getattr(config, "IMAGES_JSON_STOCK_COLUMN", "Images JSON Stock")),
    ("enhanced", getattr(config, "IMAGES_JSON_ENHANCED_COLUMN", "Images JSON Enhanced")),
]

# Reuse field definitions from the exporter to stay in sync
try:
    from inventory_data_collector import CATEGORY_COLUMNS, get_important_fields  # type: ignore
except Exception:
    # Fallback: if import fails, define empty mapping to avoid crashes and rely on config columns
    CATEGORY_COLUMNS = {}
    def get_important_fields() -> List[str]:  # type: ignore
        return [
            getattr(config, "BUYING_ENTITY_COLUMN", "Buying Entity"),
            getattr(config, "SUPPLIER_COLUMN", "Supplier"),
            getattr(config, "INVOICE_COLUMN", "Invoice"),
            getattr(config, "INVOICE_DATE_COLUMN", "Invoice Date"),
            getattr(config, "PRICE_NET_COLUMN", "Price Net"),
            getattr(config, "SHIPPING_NET_COLUMN", "Shipping Net"),
            getattr(config, "TOTAL_COST_NET_COLUMN", "Total Cost Net"),
            getattr(config, "OP_COLUMN", "OP"),
            getattr(config, "SUPPLIER_NUMBER_COLUMN", "Supplier Number"),
            getattr(config, "ISIN_COLUMN", "ISIN"),
            getattr(config, "TITLE_COLUMN", "Title"),
            getattr(config, "CATEGORY_COLUMN", "Ebay Category"),
            getattr(config, "EAN_COLUMN", "EAN"),
            getattr(config, "CONDITION_COLUMN", "Condition"),
            getattr(config, "GENDER_COLUMN", "Gender"),
            getattr(config, "BRAND_COLUMN", "Brand"),
            getattr(config, "COLOR_COLUMN", "Color"),
            getattr(config, "SIZE_COLUMN", "Size"),
            getattr(config, "LAGER_COLUMN", "Lager"),
            getattr(config, "MORE_DETAILS_COLUMN", "More details"),
            getattr(config, "KEYWORDS_COLUMN", "Keywords"),
            getattr(config, "MATERIALS_COLUMN", "Materials"),
        ]


@dataclass
class UpdateResult:
    sku: str
    updated_cells: List[str]
    appended: bool = False


def load_json_products(products_dir: Path) -> Dict[str, Dict[str, Any]]:
    """Load all per-SKU JSON files into a mapping {sku: flat_fields}.

    The exporter groups fields under human-readable category keys; we flatten those
    back to a {column_name: value} dict here.
    """
    result: Dict[str, Dict[str, Any]] = {}
    for path in products_dir.glob("*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"Skipping {path.name}: invalid JSON ({exc})")
            continue
        if not isinstance(data, dict) or len(data) != 1:
            print(f"Skipping {path.name}: unexpected JSON structure")
            continue
        sku = next(iter(data))
        payload = data[sku]
        if not isinstance(payload, dict):
            print(f"Skipping {path.name}: payload not a dict")
            continue

        flat: Dict[str, Any] = {}
        flat.update(extract_image_counts(payload))
        # Walk categories and copy known columns only
        for category, section in payload.items():
            if not isinstance(section, dict):
                continue
            for col, val in section.items():
                # Keep only scalar values that correspond to columns in inventory
                if isinstance(val, (str, int, float)) or val is None:
                    flat[col] = val
        result[sku] = flat
    return result


def extract_image_counts(payload: Dict[str, Any]) -> Dict[str, Optional[int]]:
    """Derive image counts from Images lists or summary counts.

    Prefers list lengths (phone/stock/enhanced) and falls back to Images.summary.
    Missing keys are skipped entirely to keep blanks in Excel.
    """
    counts: Dict[str, Optional[int]] = {}
    images_section = None
    if isinstance(payload.get("Images"), dict):
        images_section = payload["Images"]
    elif isinstance(payload.get("images"), dict):
        images_section = payload["images"]

    if not isinstance(images_section, dict):
        return counts

    summary = images_section.get("summary") if isinstance(images_section.get("summary"), dict) else None

    def _len_list(key: str) -> Optional[int]:
        val = images_section.get(key)
        if isinstance(val, list):
            return len(val)
        return None

    def _summary_count(key: str) -> Optional[int]:
        if not summary:
            return None
        val = summary.get(key)
        if isinstance(val, (int, float)):
            return int(val)
        return None

    for key, column_name in IMAGE_COUNT_SOURCES:
        list_len = _len_list(key)
        count = list_len if list_len is not None else _summary_count(key)
        if count is None:
            continue
        counts[column_name] = int(count)

    return counts


def normalize_for_compare(value: Any) -> Any:
    """Normalize values before comparison to avoid spurious updates."""
    if value is None:
        return ""
    # Convert timestamps/strings representing dates to date strings for compare
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        # Round to 2 decimals to match exporter behavior
        return round(value, 2)
    if isinstance(value, str):
        return value.strip()
    return value


def parse_possible_date(text: Any) -> Any:
    """Convert ISO-like strings to datetime for Excel cells; otherwise return original."""
    if isinstance(text, str):
        try:
            dt = pd.to_datetime(text, errors="raise")
            return dt.to_pydatetime()
        except Exception:
            return text
    return text


def find_header_row(ws: Worksheet, required_cols: Iterable[str]) -> int:
    """Locate the header row by scanning the first 20 rows for a row containing SKU and several required columns."""
    required = set(required_cols)
    # Ensure SKU is always required
    required.add(getattr(config, "SKU_COLUMN"))
    max_scan = min(20, ws.max_row)
    best_row = 0
    best_hits = 0
    for r in range(1, max_scan + 1):
        values = [c.value if c.value is not None else "" for c in ws[r]]
        headers = [str(v).strip() for v in values]
        hits = len(set(headers) & required)
        if hits > best_hits:
            best_row = r
            best_hits = hits
        if hits >= max(3, len(required) // 2):
            return r
    if best_row:
        return best_row
    raise RuntimeError("Could not locate header row with expected columns.")


def build_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Map column name -> 1-based column index from the header row."""
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = str(cell.value).strip() if cell.value is not None else ""
        if key:
            mapping[key] = idx
    return mapping


def scan_sku_rows(ws: Worksheet, header_row: int, sku_col_index: int) -> Dict[str, int]:
    """Return mapping SKU -> row index for all rows under the header until the last non-empty row."""
    mapping: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=sku_col_index).value
        if val is None or str(val).strip() == "":
            # Do not stop at first empty; tables might have gaps. Keep scanning.
            continue
        sku = str(val).strip()
        mapping[sku] = r
    return mapping


def apply_updates(
    ws: Worksheet,
    header_map: Dict[str, int],
    row_index: int,
    updates: Dict[str, Any],
) -> List[str]:
    """Apply cell updates to a row and return a list of changed column names."""
    changed: List[str] = []
    for col_name, new_val in updates.items():
        if col_name not in header_map:
            continue  # column not present in sheet
        col_idx = header_map[col_name]
        cell = ws.cell(row=row_index, column=col_idx)
        old = normalize_for_compare(cell.value)
        new = normalize_for_compare(new_val)
        if old == new:
            continue
        # Convert date-like strings back to datetime for Excel
        to_write = parse_possible_date(new_val)
        cell.value = to_write
        changed.append(col_name)
    return changed

# --- Override: locale-aware numeric comparison and writing for specific columns ---
from typing import Set as _SetForNumbers
NUMBER_COLUMNS: _SetForNumbers[str] = set([
    getattr(config, "PRICE_NET_COLUMN", "Price Net"),
    getattr(config, "SHIPPING_NET_COLUMN", "Shipping Net"),
    getattr(config, "OP_COLUMN", "OP"),
])

def apply_updates(
    ws: Worksheet,
    header_map: Dict[str, int],
    row_index: int,
    updates: Dict[str, Any],
) -> List[str]:
    """Apply updates with locale-aware numeric handling.

    - For NUMBER_COLUMNS, compare using floats after replacing ',' with '.'
      and write numeric values with a standard Excel number format.
    - For other columns, fallback to the default normalization.
    """
    changed: List[str] = []
    for col_name, new_val in updates.items():
        if col_name not in header_map:
            continue
        col_idx = header_map[col_name]
        cell = ws.cell(row=row_index, column=col_idx)

        if col_name in NUMBER_COLUMNS:
            def _to_float(v: Any) -> Optional[float]:
                if v in (None, ""):
                    return None
                try:
                    return float(str(v).strip().replace(',', '.'))
                except Exception:
                    return None
            old_num = _to_float(cell.value)
            new_num = _to_float(new_val)
            if old_num is not None and new_num is not None and round(old_num, 2) == round(new_num, 2):
                continue
            if new_num is None and old_num is None:
                continue
            # Write numeric with standard format; Excel will display with locale comma
            if new_num is None:
                cell.value = None
            else:
                cell.value = round(new_num, 2)
                try:
                    from openpyxl.styles import numbers as _numbers
                    cell.number_format = '0.00'
                except Exception:
                    pass
            changed.append(col_name)
            continue

        old = normalize_for_compare(cell.value)
        new = normalize_for_compare(new_val)
        if old == new:
            continue
        cell.value = parse_possible_date(new_val)
        changed.append(col_name)
    return changed


def ensure_row(
    ws: Worksheet, header_row: int, header_map: Dict[str, int], sku: str
) -> int:
    """Return the row index for SKU, appending a new row if necessary."""
    sku_idx = header_map[getattr(config, "SKU_COLUMN")]
    existing = scan_sku_rows(ws, header_row, sku_idx)
    if sku in existing:
        return existing[sku]
    # Append at the end
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=sku_idx).value = sku
    return new_row


def main() -> None:
    parser = argparse.ArgumentParser(description="Update inventory sheet from per-SKU JSONs.")
    parser.add_argument("--products-dir", type=Path, default=ROOT.parent / "products")
    parser.add_argument("--append-missing", action="store_true", help="Append rows for SKUs not present in the sheet")
    args = parser.parse_args()

    products_dir: Path = args.products_dir
    if not products_dir.exists():
        print(f"Products directory not found: {products_dir}")
        sys.exit(1)

    # Load incoming data
    incoming = load_json_products(products_dir)
    if not incoming:
        print("No product JSON files found to import.")
        sys.exit(0)

    # Open workbook and target sheet
    wb = load_workbook(filename=config.INVENTORY_FILE_PATH)
    if config.INVENTORY_SHEET_NAME not in wb.sheetnames:
        print(f"Sheet not found: {config.INVENTORY_SHEET_NAME}")
        sys.exit(1)
    ws = wb[config.INVENTORY_SHEET_NAME]

    # Map headers and SKU rows
    important = get_important_fields()
    image_columns = [
        getattr(config, "IMAGES_JSON_PHONE_COLUMN", "Images JSON Phone"),
        getattr(config, "IMAGES_JSON_STOCK_COLUMN", "Images JSON Stock"),
        getattr(config, "IMAGES_JSON_ENHANCED_COLUMN", "Images JSON Enhanced"),
    ]
    for col in image_columns:
        if col not in important:
            important.append(col)
    header_row = find_header_row(ws, important)
    header_map = build_header_map(ws, header_row)
    if getattr(config, "SKU_COLUMN") not in header_map:
        raise RuntimeError("SKU column not found in header map")

    # Build a filter of which columns we consider for updates (intersection of important & present)
    update_columns = [c for c in important if c in header_map]

    results: List[UpdateResult] = []
    for sku, flat in incoming.items():
        # Restrict to known update columns
        # Exclude Total Cost Net from updates (Excel formula)
        updates = {k: flat.get(k, None) for k in update_columns if k in flat and k != getattr(config, "TOTAL_COST_NET_COLUMN", "Total Cost Net")}
        # Mark presence of JSON for this SKU
        updates[JSON_COLUMN] = "Yes"
        if not updates:
            continue

        # Locate or create the row
        sku_rows = scan_sku_rows(ws, header_row, header_map[getattr(config, "SKU_COLUMN")])
        row_index = sku_rows.get(sku)
        appended = False
        if row_index is None:
            if not args.append_missing:
                print(f"SKU not in sheet (skipped): {sku}")
                continue
            row_index = ensure_row(ws, header_row, header_map, sku)
            appended = True

        changed_cols = apply_updates(ws, header_map, row_index, updates)
        if changed_cols or appended:
            results.append(UpdateResult(sku=sku, updated_cells=changed_cols, appended=appended))

    if results:
        wb.save(config.INVENTORY_FILE_PATH)

    # Short summary output
    updated = sum(1 for r in results if r.updated_cells)
    appended = sum(1 for r in results if r.appended)
    print(f"Processed: {len(incoming)} SKUs | Updated: {updated} | Appended: {appended}")

    # Optionally: no logging or printing to avoid storing logs in project
    # (removed logging per user request)

if __name__ == "__main__":
    main()



