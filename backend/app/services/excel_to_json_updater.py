"""Update per-SKU JSON files from Excel inventory sheet.

Reads Category, Status, and Lager columns from Excel and updates corresponding JSON files.
Also looks up and updates eBay Category ID when Category is updated.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

LEGACY = Path(__file__).resolve().parents[2] / "legacy"
import sys
sys.path.insert(0, str(LEGACY))
import config  # type: ignore

# Import category mapping function
from app.services.json_generation import get_category_id_for_path


@dataclass
class UpdateResult:
    sku: str
    updated: bool
    fields_changed: List[str]
    error: Optional[str] = None


def _find_header_row(ws: Worksheet) -> int:
    """Locate the header row by scanning for SKU column."""
    sku_column = getattr(config, "SKU_COLUMN")
    max_scan = min(20, ws.max_row)
    for r in range(1, max_scan + 1):
        values = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if sku_column in values:
            return r
    raise RuntimeError("Could not locate header row with SKU column.")


def _build_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Map column name -> 1-based column index."""
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = str(cell.value).strip() if cell.value is not None else ""
        if key:
            mapping[key] = idx
    return mapping


def _get_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get cell value, converting to string and stripping whitespace."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        return val.strip() or None
    return val


def update_jsons_from_excel(skus: Optional[List[str]] = None) -> Dict[str, Any]:
    """Update JSON files from Excel for Category, Status, and Lager columns.
    
    Args:
        skus: Optional list of SKUs to update. If None, updates all SKUs.
    
    Returns:
        Dict with success status, counts, and results.
    """
    products_dir = Path(getattr(config, "PRODUCTS_FOLDER_PATH"))
    if not products_dir.exists():
        return {"success": False, "message": f"Products directory not found: {products_dir}"}

    # Load Excel
    wb = load_workbook(filename=config.INVENTORY_FILE_PATH, data_only=True)
    if config.INVENTORY_SHEET_NAME not in wb.sheetnames:
        return {"success": False, "message": f"Sheet not found: {config.INVENTORY_SHEET_NAME}"}
    ws = wb[config.INVENTORY_SHEET_NAME]

    # Find headers
    header_row = _find_header_row(ws)
    header_map = _build_header_map(ws, header_row)

    # Required columns
    sku_col = getattr(config, "SKU_COLUMN")
    category_col = getattr(config, "CATEGORY_COLUMN")
    status_col = getattr(config, "STATUS_COLUMN")
    lager_col = getattr(config, "LAGER_COLUMN")

    if sku_col not in header_map:
        return {"success": False, "message": "SKU column not found in Excel"}

    # Get column indices
    sku_idx = header_map[sku_col]
    category_idx = header_map.get(category_col)
    status_idx = header_map.get(status_col)
    lager_idx = header_map.get(lager_col)

    # Filter SKUs if provided
    sku_filter = {s.strip() for s in skus} if skus else None

    results: List[UpdateResult] = []
    
    # Iterate through rows
    for row_num in range(header_row + 1, ws.max_row + 1):
        sku = _get_cell_value(ws, row_num, sku_idx)
        if not sku:
            continue
        
        sku = str(sku).strip()
        if sku_filter and sku not in sku_filter:
            continue

        # Get values from Excel
        category_val = _get_cell_value(ws, row_num, category_idx) if category_idx else None
        status_val = _get_cell_value(ws, row_num, status_idx) if status_idx else None
        lager_val = _get_cell_value(ws, row_num, lager_idx) if lager_idx else None

        # Load JSON file
        json_path = products_dir / f"{sku}.json"
        if not json_path.exists():
            results.append(UpdateResult(
                sku=sku,
                updated=False,
                fields_changed=[],
                error="JSON file not found"
            ))
            continue

        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            if sku not in data or not isinstance(data[sku], dict):
                results.append(UpdateResult(
                    sku=sku,
                    updated=False,
                    fields_changed=[],
                    error="Invalid JSON structure"
                ))
                continue

            payload = data[sku]
            fields_changed = []

            # Update Category in "Ebay Category" section
            if category_idx:
                if "Ebay Category" not in payload:
                    payload["Ebay Category"] = {}
                if payload["Ebay Category"].get("Category") != category_val:
                    payload["Ebay Category"]["Category"] = category_val
                    fields_changed.append("Category")
                    
                    # Also look up and update eBay Category ID
                    if category_val:
                        category_id = get_category_id_for_path(category_val)
                        if category_id:
                            payload["Ebay Category"]["eBay Category ID"] = category_id
                            if "eBay Category ID" not in fields_changed:
                                fields_changed.append("eBay Category ID")

            # Update Status in "Status" section
            if status_idx:
                if "Status" not in payload:
                    payload["Status"] = {}
                if payload["Status"].get("Status") != status_val:
                    payload["Status"]["Status"] = status_val
                    fields_changed.append("Status")

            # Update Lager in "Warehouse" section
            if lager_idx:
                if "Warehouse" not in payload:
                    payload["Warehouse"] = {}
                if payload["Warehouse"].get("Lager") != lager_val:
                    payload["Warehouse"]["Lager"] = lager_val
                    fields_changed.append("Lager")

            # Save if changed
            if fields_changed:
                json_path.write_text(
                    json.dumps(data, indent=2, ensure_ascii=False),
                    encoding="utf-8"
                )
                results.append(UpdateResult(
                    sku=sku,
                    updated=True,
                    fields_changed=fields_changed
                ))
            else:
                results.append(UpdateResult(
                    sku=sku,
                    updated=False,
                    fields_changed=[]
                ))

        except Exception as e:
            results.append(UpdateResult(
                sku=sku,
                updated=False,
                fields_changed=[],
                error=str(e)
            ))

    updated_count = sum(1 for r in results if r.updated)
    return {
        "success": True,
        "processed": len(results),
        "updated": updated_count,
        "message": f"Processed {len(results)} SKUs | Updated {updated_count} JSONs",
    }
