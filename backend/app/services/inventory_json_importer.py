"""Import per-SKU JSON files into the inventory Excel sheet (table-aware).

Updates only changed cells in-place with openpyxl to preserve table formatting.
Optionally appends missing SKUs and updates folder image counters.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.folder_images_cache import get_folder_image_count

LEGACY = Path(__file__).resolve().parents[2] / "legacy"
import sys
sys.path.insert(0, str(LEGACY))
import config  # type: ignore


JSON_COLUMN = getattr(config, "JSON_COLUMN", "JSON")
IMAGES_COLUMN = getattr(config, "IMAGES_COLUMN", "Images")
FOLDER_IMAGES_COLUMN = getattr(config, "FOLDER_IMAGES_COLUMN", "Folder Images")

IMAGE_COUNT_SOURCES = [
    ("phone", getattr(config, "IMAGES_JSON_PHONE_COLUMN", "Images JSON Phone")),
    ("stock", getattr(config, "IMAGES_JSON_STOCK_COLUMN", "Images JSON Stock")),
    ("enhanced", getattr(config, "IMAGES_JSON_ENHANCED_COLUMN", "Images JSON Enhanced")),
]


@dataclass
class UpdateResult:
    sku: str
    updated_cells: List[str]
    appended: bool = False


def _load_json_products(products_dir: Path, skus: Optional[List[str]] = None) -> Dict[str, Dict[str, Any]]:
    """Load per-SKU JSON files into a mapping {sku: flat_fields}."""
    result: Dict[str, Dict[str, Any]] = {}
    sku_filter = {s.strip() for s in skus} if skus else None

    for path in products_dir.glob("*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(data, dict) or len(data) != 1:
            continue
        sku = next(iter(data)).strip()
        if sku_filter is not None and sku not in sku_filter:
            continue
        payload = data[sku]
        if not isinstance(payload, dict):
            continue

        flat: Dict[str, Any] = {}
        flat.update(_extract_image_counts(payload))

        for _, section in payload.items():
            if not isinstance(section, dict):
                continue
            for col, val in section.items():
                if isinstance(val, (str, int, float)) or val is None:
                    flat[col] = val

        result[sku] = flat
    return result


def _extract_image_counts(payload: Dict[str, Any]) -> Dict[str, Optional[int]]:
    counts: Dict[str, Optional[int]] = {}
    images_section = payload.get("Images") if isinstance(payload.get("Images"), dict) else None
    if not isinstance(images_section, dict):
        return counts

    summary = images_section.get("summary") if isinstance(images_section.get("summary"), dict) else None

    def _len_list(key: str) -> Optional[int]:
        val = images_section.get(key)
        if isinstance(val, list):
            return len(val)
        return None

    def _summary_count(key: str) -> Optional[int]:
        if not summary:
            return None
        val = summary.get(key)
        if isinstance(val, (int, float)):
            return int(val)
        return None

    for key, column_name in IMAGE_COUNT_SOURCES:
        list_len = _len_list(key)
        count = list_len if list_len is not None else _summary_count(f"count_{key}")
        if count is None:
            continue
        counts[column_name] = int(count)

    return counts


def _normalize_for_compare(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float):
        return round(value, 2)
    if isinstance(value, str):
        return value.strip()
    return value


def _parse_possible_date(text: Any) -> Any:
    if isinstance(text, str):
        try:
            dt = pd.to_datetime(text, errors="raise")
            return dt.to_pydatetime()
        except Exception:
            return text
    return text


def _find_header_row(ws: Worksheet, required_cols: Iterable[str]) -> int:
    required = set(required_cols)
    required.add(getattr(config, "SKU_COLUMN"))
    max_scan = min(20, ws.max_row)
    best_row = 0
    best_hits = 0
    for r in range(1, max_scan + 1):
        values = [c.value if c.value is not None else "" for c in ws[r]]
        headers = [str(v).strip() for v in values]
        hits = len(set(headers) & required)
        if hits > best_hits:
            best_row = r
            best_hits = hits
        if hits >= max(3, len(required) // 2):
            return r
    if best_row:
        return best_row
    raise RuntimeError("Could not locate header row with expected columns.")


def _build_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = str(cell.value).strip() if cell.value is not None else ""
        if key:
            mapping[key] = idx
    return mapping


def _scan_sku_rows(ws: Worksheet, header_row: int, sku_col_index: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=sku_col_index).value
        if val is None or str(val).strip() == "":
            continue
        sku = str(val).strip()
        mapping[sku] = r
    return mapping


def _apply_updates(
    ws: Worksheet,
    header_map: Dict[str, int],
    row_index: int,
    updates: Dict[str, Any],
) -> List[str]:
    changed: List[str] = []
    for col_name, new_val in updates.items():
        if col_name not in header_map:
            continue
        col_idx = header_map[col_name]
        cell = ws.cell(row=row_index, column=col_idx)
        old = _normalize_for_compare(cell.value)
        new = _normalize_for_compare(new_val)
        if old == new:
            continue
        cell.value = _parse_possible_date(new_val)
        changed.append(col_name)
    return changed


def _ensure_row(ws: Worksheet, header_row: int, header_map: Dict[str, int], sku: str) -> int:
    sku_idx = header_map[getattr(config, "SKU_COLUMN")]
    existing = _scan_sku_rows(ws, header_row, sku_idx)
    if sku in existing:
        return existing[sku]
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=sku_idx).value = sku
    return new_row


def import_jsons_to_inventory(
    skus: Optional[List[str]] = None,
    append_missing: bool = False,
) -> Dict[str, Any]:
    products_dir = Path(getattr(config, "PRODUCTS_FOLDER_PATH"))
    if not products_dir.exists():
        return {"success": False, "message": f"Products directory not found: {products_dir}"}

    incoming = _load_json_products(products_dir, skus=skus)
    if not incoming:
        return {"success": False, "message": "No product JSON files found to import."}

    wb = load_workbook(filename=config.INVENTORY_FILE_PATH)
    if config.INVENTORY_SHEET_NAME not in wb.sheetnames:
        return {"success": False, "message": f"Sheet not found: {config.INVENTORY_SHEET_NAME}"}
    ws = wb[config.INVENTORY_SHEET_NAME]

    # Build header map
    required_cols = list(incoming[next(iter(incoming))].keys())
    required_cols.extend([JSON_COLUMN, IMAGES_COLUMN, FOLDER_IMAGES_COLUMN])
    
    # Explicitly add image count columns to ensure they're found in header
    for _, column_name in IMAGE_COUNT_SOURCES:
        if column_name not in required_cols:
            required_cols.append(column_name)
    
    header_row = _find_header_row(ws, required_cols)
    header_map = _build_header_map(ws, header_row)
    if getattr(config, "SKU_COLUMN") not in header_map:
        return {"success": False, "message": "SKU column not found in header map"}

    update_columns = [c for c in required_cols if c in header_map]

    results: List[UpdateResult] = []
    for sku, flat in incoming.items():
        updates = {k: flat.get(k, None) for k in update_columns if k in flat}
        updates[JSON_COLUMN] = "Yes"

        # Add folder image counts
        folder_count = get_folder_image_count(sku)
        if folder_count is not None:
            if FOLDER_IMAGES_COLUMN in header_map:
                updates[FOLDER_IMAGES_COLUMN] = int(folder_count)
            if IMAGES_COLUMN in header_map:
                updates[IMAGES_COLUMN] = int(folder_count)

        if not updates:
            continue

        sku_rows = _scan_sku_rows(ws, header_row, header_map[getattr(config, "SKU_COLUMN")])
        row_index = sku_rows.get(sku)
        appended = False
        if row_index is None:
            if not append_missing:
                continue
            row_index = _ensure_row(ws, header_row, header_map, sku)
            appended = True

        changed_cols = _apply_updates(ws, header_map, row_index, updates)
        if changed_cols or appended:
            results.append(UpdateResult(sku=sku, updated_cells=changed_cols, appended=appended))

    if results:
        wb.save(config.INVENTORY_FILE_PATH)

    updated = sum(1 for r in results if r.updated_cells)
    appended = sum(1 for r in results if r.appended)
    return {
        "success": True,
        "processed": len(incoming),
        "updated": updated,
        "appended": appended,
        "message": f"Processed {len(incoming)} SKUs | Updated {updated} | Appended {appended}",
    }