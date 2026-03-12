"""Service to selectively sync columns from Excel to SQLite database."""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from app.services.excel_inventory import excel_inventory, _get_db_path

LEGACY = Path(__file__).resolve().parents[2] / "legacy"
import sys
sys.path.insert(0, str(LEGACY))
import config  # type: ignore


def get_excel_sheets() -> List[str]:
    """Get list of sheet names from Excel file."""
    try:
        wb = load_workbook(config.INVENTORY_FILE_PATH)
        return wb.sheetnames
    except Exception as e:
        print(f"Error reading Excel sheets: {e}")
        return []


def get_excel_columns(sheet_name: str) -> List[str]:
    """Get list of column names from a specific Excel sheet."""
    try:
        df = pd.read_excel(config.INVENTORY_FILE_PATH, sheet_name=sheet_name, nrows=0)
        return list(df.columns)
    except Exception as e:
        print(f"Error reading columns from sheet {sheet_name}: {e}")
        return []


def _find_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    alias_set = {a.strip().lower() for a in aliases}
    for col in columns:
        if str(col).strip().lower() in alias_set:
            return col
    return None


def _to_text_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_optional_float(value: Any) -> Optional[float]:
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except Exception:
        return None


def refresh_category_mapping_from_excel(sheet_name: str = "Ebay Categories") -> Dict[str, Any]:
    """Update backend/schemas/category_mapping.json from Excel category sheet.

    Existing entries are updated by categoryId (preferred) or fullPath, and missing
    entries are appended. This keeps existing mappings while refreshing fee data.
    """
    try:
        df = pd.read_excel(config.INVENTORY_FILE_PATH, sheet_name=sheet_name)
        if df.empty:
            return {"success": False, "message": f"Sheet '{sheet_name}' is empty"}

        columns = list(df.columns)
        category_col = _find_column(columns, ["Category", "Category Path", "fullPath"])
        id_col = _find_column(columns, ["ID", "Category ID", "CategoryId"])
        payment_fee_col = _find_column(columns, ["Payment Fee", "Payment Fee EUR", "Payment Fee €", "payment_fee"])
        up_to_amount_col = _find_column(columns, ["Up To", "Final Amount Up To", "final_amount_up_to"])
        commission_up_to_col = _find_column(columns, ["Sales commission per item Up To", "Sales Commission Up To", "sales_commission_up_to"])
        from_amount_col = _find_column(columns, ["From", "Final Amount From", "final_amount_from"])
        commission_from_col = _find_column(columns, ["Sales commission per item From", "Sales Commission From", "sales_commission_from"])

        if not category_col or not id_col:
            return {
                "success": False,
                "message": "Required columns not found in Excel sheet. Need at least 'Category' and 'ID'.",
            }

        mapping_path = Path(__file__).resolve().parents[2] / "schemas" / "category_mapping.json"
        existing_payload: Dict[str, Any] = {"categoryMappings": []}
        if mapping_path.exists():
            with open(mapping_path, "r", encoding="utf-8") as f:
                existing_payload = json.load(f) or {"categoryMappings": []}

        existing_mappings = existing_payload.get("categoryMappings") or []
        if not isinstance(existing_mappings, list):
            existing_mappings = []

        by_id: Dict[str, Dict[str, Any]] = {}
        by_path: Dict[str, Dict[str, Any]] = {}
        for entry in existing_mappings:
            if not isinstance(entry, dict):
                continue
            entry_id = _to_text_id(entry.get("categoryId"))
            entry_path = str(entry.get("fullPath") or "").strip().lower()
            if entry_id:
                by_id[entry_id] = entry
            if entry_path:
                by_path[entry_path] = entry

        rows_processed = 0
        rows_skipped = 0
        rows_updated = 0
        rows_added = 0

        for _, row in df.iterrows():
            category_path = str(row.get(category_col) or "").strip()
            category_id = _to_text_id(row.get(id_col))
            if not category_path or not category_id:
                rows_skipped += 1
                continue

            rows_processed += 1
            category_name = category_path.split("/")[-1].strip() if "/" in category_path else category_path
            fees = {
                "payment_fee": _to_optional_float(row.get(payment_fee_col)) if payment_fee_col else None,
                "final_amount_up_to": _to_optional_float(row.get(up_to_amount_col)) if up_to_amount_col else None,
                "sales_commission_up_to": _to_optional_float(row.get(commission_up_to_col)) if commission_up_to_col else None,
                "final_amount_from": _to_optional_float(row.get(from_amount_col)) if from_amount_col else None,
                "sales_commission_from": _to_optional_float(row.get(commission_from_col)) if commission_from_col else None,
            }

            target = by_id.get(category_id) or by_path.get(category_path.lower())
            if target is not None:
                target["categoryId"] = category_id
                target["fullPath"] = category_path
                target["categoryName"] = category_name
                target["market"] = target.get("market") or "EBAY_DE"
                target["fees"] = fees
                rows_updated += 1
            else:
                new_entry = {
                    "categoryName": category_name,
                    "categoryId": category_id,
                    "fullPath": category_path,
                    "market": "EBAY_DE",
                    "fees": fees,
                }
                existing_mappings.append(new_entry)
                by_id[category_id] = new_entry
                by_path[category_path.lower()] = new_entry
                rows_added += 1

        mapping_path.parent.mkdir(parents=True, exist_ok=True)
        with open(mapping_path, "w", encoding="utf-8") as f:
            json.dump({"categoryMappings": existing_mappings}, f, ensure_ascii=False, indent=2)

        return {
            "success": True,
            "message": "category_mapping.json refreshed from Excel",
            "sheet": sheet_name,
            "rows_processed": rows_processed,
            "rows_skipped": rows_skipped,
            "rows_updated": rows_updated,
            "rows_added": rows_added,
            "total_mappings": len(existing_mappings),
            "file": str(mapping_path),
        }
    except Exception as e:
        print(f"[CATEGORY MAPPING] Failed to refresh from Excel: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"Error: {str(e)}"}


def sync_excel_to_db(sheet_name: str, columns: List[str]) -> Dict[str, Any]:
    """Sync specific columns from an Excel sheet to the corresponding database table.
    Matches rows by unique identifiers, updates existing rows and inserts new ones.
    
    Args:
        sheet_name: Name of the Excel sheet
        columns: List of column names to sync
    
    Returns:
        Dict with success, message, and stats
    """
    if not sheet_name or not columns:
        return {"success": False, "message": "Sheet name and columns are required"}

    # Map sheet names to table names (lowercase with underscores)
    table_name = sheet_name.lower().replace(" ", "_")

    # Define unique key columns for each sheet
    UNIQUE_KEYS = {
        "inventory": ["SKU (Old)"],
        "listings": ["SKU (Old)"],
        "expenses": ["Pay Date", "Invoice", "Partner"],  # Combo to identify an expense
        "income": ["Pay Date", "Platform", "Item Title", "Sale Date", "SKU (Old)"],  # Combo to identify a sale
        "ebay_categories": ["Category", "ID"],  # Category name + ID
    }

    # Tables that should be completely replaced (delete all, insert all)
    FULL_REPLACE_TABLES = ["income"]

    try:
        # Read Excel sheet
        df = pd.read_excel(config.INVENTORY_FILE_PATH, sheet_name=sheet_name)
        if df.empty:
            return {"success": False, "message": f"Sheet '{sheet_name}' is empty"}

        # Filter to only requested columns that exist
        available_cols = [c for c in columns if c in df.columns]
        if not available_cols:
            return {"success": False, "message": f"None of the requested columns exist in sheet '{sheet_name}'"}

        # Connect to database
        db_path = _get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        try:
            # Check if table exists
            table_check = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (table_name,)
            ).fetchone()
            
            if not table_check:
                return {"success": False, "message": f"Table '{table_name}' not found in database"}

            # Get all columns in the table
            table_info = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
            db_columns = {r[1]: r[0] for r in table_info}  # {col_name: col_id}

            if not db_columns:
                return {"success": False, "message": f"Table '{table_name}' has no columns"}

            # Get unique key columns for this table
            unique_keys = UNIQUE_KEYS.get(table_name, [])
            if not unique_keys:
                return {"success": False, "message": f"No unique key defined for table '{table_name}'. Cannot safely sync."}

            # Verify all unique key columns exist in both Excel and DB
            missing_in_excel = [k for k in unique_keys if k not in df.columns]
            missing_in_db = [k for k in unique_keys if k not in db_columns]
            if missing_in_excel:
                return {"success": False, "message": f"Unique key columns missing in Excel: {missing_in_excel}"}
            if missing_in_db:
                return {"success": False, "message": f"Unique key columns missing in DB: {missing_in_db}"}

            print(f"[SYNC] Sheet: {sheet_name}, Table: {table_name}, Unique keys: {unique_keys}, Syncing cols: {available_cols}")

            rows_updated = 0
            rows_inserted = 0
            rows_failed = 0
            rows_skipped_insert = 0

            # Check if this table should be fully replaced
            if table_name in FULL_REPLACE_TABLES:
                print(f"[SYNC] {table_name} is marked for FULL REPLACEMENT - deleting all existing rows...")
                conn.execute(f'DELETE FROM "{table_name}"')
                conn.commit()

                # INSERT all rows fresh from Excel
                for idx, (_, row) in enumerate(df.iterrows()):
                    try:
                        all_values = {}
                        for db_col in db_columns.keys():
                            if db_col in df.columns:
                                val = row[db_col]
                                if pd.isna(val):
                                    val = None
                                elif hasattr(val, 'isoformat'):
                                    val = val.isoformat()
                                all_values[db_col] = val
                            else:
                                all_values[db_col] = None
                        
                        cols_str = ", ".join([f'"{col}"' for col in all_values.keys()])
                        placeholders = ", ".join(['?' for _ in all_values])
                        conn.execute(
                            f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})',
                            list(all_values.values()),
                        )
                        rows_inserted += 1

                    except Exception as e:
                        print(f"[SYNC ERROR] Row {idx}: {e}")
                        import traceback
                        traceback.print_exc()
                        rows_failed += 1

                conn.commit()
                print(f"[SYNC RESULT] Full replacement: inserted {rows_inserted} rows, failed {rows_failed}")
            
            else:
                # Smart sync: Load all existing rows from DB and match by unique key
                db_rows = {}
                all_db_rows = conn.execute(f'SELECT rowid, * FROM "{table_name}"').fetchall()
                for db_row in all_db_rows:
                    # Build unique key tuple
                    key_values = []
                    for key_col in unique_keys:
                        val = db_row[key_col]
                        # Normalize for comparison
                        if isinstance(val, str):
                            val = val.strip()
                        key_values.append(val)
                    key_tuple = tuple(key_values)
                    db_rows.setdefault(key_tuple, []).append(db_row["rowid"])

                print(f"[SYNC] Found {len(db_rows)} existing rows in database")

                # Sync rows - UPDATE existing by unique key, INSERT new
                for idx, (_, row) in enumerate(df.iterrows()):
                    try:
                        # Build unique key for this Excel row
                        key_values = []
                        for key_col in unique_keys:
                            val = row[key_col]
                            if pd.isna(val):
                                val = None
                            elif hasattr(val, 'isoformat'):
                                val = val.isoformat()
                            elif isinstance(val, str):
                                val = val.strip()
                            key_values.append(val)
                        key_tuple = tuple(key_values)

                        # Skip if unique key has None values
                        if None in key_values:
                            rows_failed += 1
                            continue

                        # Check if this row exists in database
                        existing_row_ids = db_rows.get(key_tuple)

                        if existing_row_ids:
                            # UPDATE existing row
                            updates = {}
                            for col in available_cols:
                                if col not in db_columns:
                                    continue
                                val = row[col]
                                if pd.isna(val):
                                    val = None
                                elif hasattr(val, 'isoformat'):
                                    val = val.isoformat()
                                updates[col] = val

                            if updates:
                                set_clause = ", ".join([f'"{col}" = ?' for col in updates.keys()])
                                for rowid in existing_row_ids:
                                    conn.execute(
                                        f'UPDATE "{table_name}" SET {set_clause} WHERE rowid = ?',
                                        list(updates.values()) + [rowid],
                                    )
                                rows_updated += len(existing_row_ids)
                        else:
                            # INSERT new row - include ALL columns from Excel
                            if table_name == "inventory":
                                rows_skipped_insert += 1
                                continue
                            all_values = {}
                            for db_col in db_columns.keys():
                                if db_col in df.columns:
                                    val = row[db_col]
                                    if pd.isna(val):
                                        val = None
                                    elif hasattr(val, 'isoformat'):
                                        val = val.isoformat()
                                    all_values[db_col] = val
                                else:
                                    all_values[db_col] = None
                            
                            cols_str = ", ".join([f'"{col}"' for col in all_values.keys()])
                            placeholders = ", ".join(['?' for _ in all_values])
                            conn.execute(
                                f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})',
                                list(all_values.values()),
                            )
                            rows_inserted += 1

                    except Exception as e:
                        print(f"[SYNC ERROR] Row {idx}: {e}")
                        import traceback
                        traceback.print_exc()
                        rows_failed += 1

                # Commit updates/inserts for smart sync
                conn.commit()

            # Invalidate cache
            if table_name == "inventory":
                excel_inventory.invalidate()

            return {
                "success": True,
                "message": f"Synced '{sheet_name}': updated {rows_updated} rows, inserted {rows_inserted} new rows",
                "rows_updated": rows_updated,
                "rows_inserted": rows_inserted,
                "rows_failed": rows_failed,
                "rows_skipped_insert": rows_skipped_insert,
                "columns_synced": available_cols,
            }
        finally:
            conn.close()

    except Exception as e:
        print(f"[SYNC EXCEPTION] Error syncing Excel to DB: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"Error: {str(e)}"}


def add_missing_sku_rows_from_excel() -> Dict[str, Any]:
    """Insert only missing SKU rows from Excel Inventory sheet into DB inventory table.

    Matching key: "SKU (Old)"
    - Existing SKUs in DB are not modified.
    - Missing SKUs are inserted with all available Excel columns mapped to DB columns.
    """
    sheet_name = "Inventory"
    table_name = "inventory"
    sku_col = "SKU (Old)"

    try:
        df = pd.read_excel(config.INVENTORY_FILE_PATH, sheet_name=sheet_name)
        if df.empty:
            return {"success": False, "message": f"Sheet '{sheet_name}' is empty"}
        if sku_col not in df.columns:
            return {"success": False, "message": f"Column '{sku_col}' not found in sheet '{sheet_name}'"}

        db_path = _get_db_path()
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        try:
            table_check = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (table_name,)
            ).fetchone()
            if not table_check:
                return {"success": False, "message": f"Table '{table_name}' not found in database"}

            table_info = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
            db_columns = [r[1] for r in table_info]
            db_columns_set = set(db_columns)
            if sku_col not in db_columns_set:
                return {"success": False, "message": f"Column '{sku_col}' not found in DB table '{table_name}'"}

            existing_rows = conn.execute(f'SELECT "{sku_col}" FROM "{table_name}"').fetchall()
            existing_skus = {
                str(row[sku_col]).strip()
                for row in existing_rows
                if row[sku_col] is not None and str(row[sku_col]).strip()
            }

            rows_inserted = 0
            rows_failed = 0
            rows_existing = 0
            rows_invalid_sku = 0

            for idx, (_, row) in enumerate(df.iterrows()):
                raw_sku = row.get(sku_col)
                if pd.isna(raw_sku) or str(raw_sku).strip() == "":
                    rows_invalid_sku += 1
                    continue

                sku_value = str(raw_sku).strip()
                if sku_value in existing_skus:
                    rows_existing += 1
                    continue

                try:
                    insert_values: Dict[str, Any] = {}
                    for db_col in db_columns:
                        if db_col in df.columns:
                            val = row[db_col]
                            if pd.isna(val):
                                val = None
                            elif hasattr(val, "isoformat"):
                                val = val.isoformat()
                            insert_values[db_col] = val
                        else:
                            insert_values[db_col] = None

                    cols_str = ", ".join([f'"{col}"' for col in insert_values.keys()])
                    placeholders = ", ".join(["?" for _ in insert_values])
                    conn.execute(
                        f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})',
                        list(insert_values.values())
                    )

                    existing_skus.add(sku_value)
                    rows_inserted += 1
                except Exception as e:
                    print(f"[ADD MISSING SKU] Row {idx} failed: {e}")
                    rows_failed += 1

            conn.commit()
            excel_inventory.invalidate()

            return {
                "success": True,
                "message": f"Added {rows_inserted} missing SKU row(s) from Excel to DB",
                "rows_inserted": rows_inserted,
                "rows_existing": rows_existing,
                "rows_invalid_sku": rows_invalid_sku,
                "rows_failed": rows_failed,
                "sheet": sheet_name,
                "table": table_name,
            }
        finally:
            conn.close()
    except Exception as e:
        print(f"[ADD MISSING SKU] Error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"Error: {str(e)}"}
