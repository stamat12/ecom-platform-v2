"""Service to sync JSON data back to Excel."""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

LEGACY = Path(__file__).resolve().parents[2] / "legacy"
import sys
sys.path.insert(0, str(LEGACY))
import config  # type: ignore
from app.services.excel_inventory import _get_db_path

# Excel columns to sync from JSON
EXCEL_COLUMNS_TO_SYNC = [
    "Category",           # from Ebay Category section
    "EAN",                # from EAN section
    "Condition",          # from Product Condition section
    "Gender",             # from Intern Product Info section
    "Brand",              # from Intern Product Info section
    "Color",              # from Intern Product Info section
    "Size",               # from Intern Product Info section
    "More details",       # from Intern Generated Info section
    "Materials",          # from Intern Generated Info section
    "Keywords",           # from Intern Generated Info section
    "OP",                 # from OP section
    "Status",             # from Status section
    "Lager",              # from Warehouse section
    "Images JSON Phone",  # from Images section (count)
    "Images JSON Stock",  # from Images section (count)
    "Images JSON Enhanced",  # from Images section (count)
    "JSON",               # whether JSON file exists (Yes/empty)
    "Images"              # folder images count from cache
]

PRODUCTS_DIR = LEGACY / "products"


def load_db_inventory() -> pd.DataFrame:
    """Load inventory table from SQLite database."""
    db_path = _get_db_path()
    conn = sqlite3.connect(db_path)
    try:
        return pd.read_sql("SELECT * FROM inventory", conn)
    finally:
        conn.close()


def detect_changes(current_values: Dict[str, Any], new_values: Dict[str, Any]) -> bool:
    """Detect if there are any changes between current and new values."""
    for key in set(list(current_values.keys()) + list(new_values.keys())):
        curr = current_values.get(key)
        new = new_values.get(key)
        
        # Normalize JSON strings for comparison
        if isinstance(curr, str) and curr and curr[0] in '{[':
            try:
                curr = json.loads(curr)
            except:
                pass
        
        if isinstance(new, str) and new and new[0] in '{[':
            try:
                new = json.loads(new)
            except:
                pass
        
        # Compare values
        if curr != new:
            return True
    
    return False


def sync_db_to_excel(sheet_name: str = "Inventory") -> Dict[str, Any]:
    """
    Sync data from JSON files back to Excel.
    Only updates cells where there are changes.
    
    Args:
        sheet_name: Name of the Excel sheet to update
    
    Returns:
        Dict with sync statistics
    """
    try:
        # Load Excel file
        excel_file = config.INVENTORY_FILE_PATH
        if not excel_file.exists():
            return {"success": False, "message": f"Excel file not found: {excel_file}"}
        
        # Read current Excel data
        df = pd.read_excel(excel_file, sheet_name=sheet_name)

        # Load inventory data from DB
        db_df = load_db_inventory()
        if db_df.empty:
            return {"success": False, "message": "Inventory DB is empty"}

        db_sku_col = "SKU (Old)" if "SKU (Old)" in db_df.columns else ("SKU" if "SKU" in db_df.columns else None)
        if not db_sku_col:
            return {"success": False, "message": "Could not find SKU column in DB"}

        db_lookup = {}
        for _, db_row in db_df.iterrows():
            sku_val = db_row.get(db_sku_col)
            if pd.isna(sku_val) or sku_val is None:
                continue
            sku_key = str(sku_val).strip()
            if not sku_key:
                continue
            db_lookup[sku_key] = db_row
        
        # Track changes
        changes_made = 0
        rows_processed = 0
        columns_updated = set()
        
        # Load workbook for direct cell editing
        wb = load_workbook(excel_file)
        ws = wb[sheet_name]
        
        # Get header row (first row)
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        
        # Columns to update
        columns_to_update = EXCEL_COLUMNS_TO_SYNC
        
        # Find rows by SKU and update
        sku_col = None
        for col_name, col_idx in headers.items():
            if col_name and "sku" in str(col_name).lower() and "old" in str(col_name).lower():
                sku_col = col_idx
                break
        
        if not sku_col:
            return {"success": False, "message": "Could not find SKU column in Excel"}
        
        # Process each row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            sku_cell = ws.cell(row=row_idx, column=sku_col)
            sku = sku_cell.value
            
            if not sku:
                continue
            
            rows_processed += 1
            
            # Get DB row for this SKU
            sku_key = str(sku).strip()
            db_row = db_lookup.get(sku_key)
            if db_row is None:
                continue
            
            # Update cells where there are changes
            for col_name in columns_to_update:
                if col_name not in headers:
                    continue
                
                col_idx = headers[col_name]
                current_cell = ws.cell(row=row_idx, column=col_idx)
                current_value = current_cell.value
                if col_name == "Status" and str(current_value).strip() == "OK":
                    continue
                if col_name not in db_df.columns:
                    continue
                new_value = db_row.get(col_name)
                if pd.isna(new_value):
                    new_value = None
                
                # Convert new value to appropriate format for Excel
                if isinstance(new_value, dict):
                    new_value = json.dumps(new_value, ensure_ascii=False)
                elif isinstance(new_value, (list, tuple)):
                    new_value = json.dumps(new_value, ensure_ascii=False)
                
                # Detect changes
                if detect_changes({col_name: current_value}, {col_name: new_value}):
                    current_cell.value = new_value
                    changes_made += 1
                    columns_updated.add(col_name)
        
        # Save workbook
        wb.save(excel_file)
        
        return {
            "success": True,
            "message": f"Successfully synced {rows_processed} rows from JSON to Excel",
            "stats": {
                "rows_processed": rows_processed,
                "changes_made": changes_made,
                "columns_updated": sorted(list(columns_updated))
            }
        }
    
    except Exception as e:
        return {
            "success": False,
            "message": f"Error during sync: {str(e)}"
        }
